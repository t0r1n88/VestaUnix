"""
Извлечение данных из файлов Excel со сложной структурой
"""
from support_functions import write_df_to_excel # импорт функции по записи в файл с автошириной колонок
import pandas as pd
from tkinter import messagebox
import openpyxl
import time
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',)

def count_text_value(df):
    """
    Функция для подсчета количества вариантов того или иного показателя
    :param df: датафрейм с сырыми данными. Название показателя значение показателя(строка разделенная ;)
    :return: обработанный датафрейм с мультиндексом, где (Название показателя это индекс верхнего уровня, вариант показателя это индекс второго уровня а значение это сколько раз встречался
    этот вариант в обрабатываемых файлах)
    """
    data = dict()

    #
    for row in df.itertuples():
        value = row[2]
        if type(value) == float or type(value) == int:
            continue
        # Создаем список, разделяя строку по ;
        lst_value = row[2].split(';')[:-1]
        #     # Отрезаем последний элемент, поскольку это пустое значение
        temp_df = pd.DataFrame({'Value': lst_value})
        counts_series = temp_df['Value'].value_counts()
        # Делаем индекс колонкой и превращаем в обычную таблицу
        index_count_values = counts_series.reset_index()
        # Итерируемся по таблице.Это делается чтобы заполнить словарь на основе которого будет создаваться итоговая таблица
        for count_row in index_count_values.itertuples():
            # print(count_row)
            # Заполняем словарь
            data[(row[1], count_row[1])] = count_row[2]
    # Создаем на основе получившегося словаря таблицу
    out_df = pd.Series(data).to_frame().reset_index()
    #TODO обработка пустых датафрейма
    out_df = out_df.set_index(['level_0', 'level_1'])
    out_df.index.names = ['Название показателя', 'Вариант показателя']
    out_df.rename(columns={0: 'Количество'}, inplace=True)
    return out_df


def check_data(cell, text_mode):
    """
    Функция для проверки значения ячейки. Для обработки пустых значений, строковых значений, дат
    :param cell: значение ячейки
    :return: 0 если значение ячейки не число
            число если значение ячейки число(ха звучит глуповато)
    думаю функция должна работать с дополнительным параметром, от которого будет зависеть подсчет значений навроде галочек или плюсов в анкетах или опросах.
    """
    # Проверяем режим работы. если текстовый, то просто складываем строки
    if text_mode == 'Yes':
        if cell is None:
            return ''
        else:
            temp_str = str(cell)
            return f'{temp_str};'
    # Если режим работы стандартный. Убрал подсчет строк и символов в числовом режиме, чтобы не запутывать.
    else:
        if cell is None:
            return 0
        if type(cell) == int:
            return cell
        elif type(cell) == float:
            return cell
        else:
            return 0


def extract_data_from_hard_xlsx(mode_text,name_file_params_calculate_data,names_files_calculate_data,path_to_end_folder_calculate_data):
    """
    Функция для извлечения данных из таблиц Excel со сложной структурой, извлечение происходит из конкретных ячеек указанных в файле параметров
    :param mode_text: режим работы (обработка текста или чисел)
    :param name_file_params_calculate_data: файл  указанием ячеек данные из которых нужно извлечь
    :param names_files_calculate_data: файлы которые нужно обработать
    :param path_to_end_folder_calculate_data:  итоговая папка
    :return:
    """
    try:
        count = 0
        count_errors = 0
        quantity_files = len(names_files_calculate_data)
        current_time = time.strftime('%H_%M_%S')
        # Состояние чекбокса

        # Получаем название обрабатываемого листа
        name_list_df = pd.read_excel(name_file_params_calculate_data, nrows=2)
        name_list = name_list_df['Значение'].loc[0]

        # Получаем количество листов в файле, на случай если название листа не совпадает с правильным
        quantity_list_in_file = name_list_df['Значение'].loc[1]

        # Получаем шаблон с данными, первую строку пропускаем, поскольку название обрабатываемого листа мы уже получили
        df = pd.read_excel(name_file_params_calculate_data, skiprows=2)

        # Создаем словарь параметров
        param_dict = dict()

        for row in df.itertuples():
            param_dict[row[1]] = row[2]
        # Создаем словарь для подсчета данных, копируя ключи из словаря параметров, значения в зависимости от способа обработки

        if mode_text == 'Yes':
            result_dct = {key: '' for key, value in param_dict.items()}
        else:
            result_dct = {key: 0 for key, value in param_dict.items()}

            # Создаем датафрейм для контроля процесса подсчета и заполняем словарь на основе которого будем делать итоговую таблицу

        check_df = pd.DataFrame(columns=param_dict.keys())
        # Вставляем колонку для названия файла
        check_df.insert(0, 'Название файла', '')
        for file in names_files_calculate_data:
            name_file = file.split('/')[-1]
            print(name_file) # обрабатываемый файл
            # Проверяем чтобы файл не был резервной копией или файлом с другим расширением.
            if file.startswith('~$'):
                continue
            # Создаем словарь для создания строки которую мы будем добавлять в проверочный датафрейм
            new_row = dict()
            # Получаем  отбрасываем расширение файла
            full_name_file = file.split('.')[0]
            # Получаем имя файла  без пути
            name_file = full_name_file.split('/')[-1]
            try:

                new_row['Название файла'] = name_file

                wb = openpyxl.load_workbook(file)
                # Проверяем наличие листа
                if name_list in wb.sheetnames:
                    sheet = wb[name_list]
                # проверяем количество листов в файле.Если значение равно 1 то просто берем первый лист, иначе вызываем ошибку
                elif quantity_list_in_file == 1:
                    temp_name = wb.sheetnames[0]
                    sheet = wb[temp_name]
                else:
                    raise Exception
                for key, cell in param_dict.items():
                    result_dct[key] += check_data(sheet[cell].value, mode_text)
                    new_row[key] = sheet[cell].value

                temp_df = pd.DataFrame(new_row, index=['temp_index'])
                check_df = pd.concat([check_df, temp_df], ignore_index=True)
                # check_df = check_df.append(new_row, ignore_index=True)

                count += 1
            # Ловим исключения
            except Exception as err:
                count_errors += 1
                with open(f'{path_to_end_folder_calculate_data}/Необработанные файлы {current_time}.txt', 'a',
                          encoding='utf-8') as f:
                    f.write(f'Файл {name_file} не обработан!!!\n')

        # сохраняем

        check_df.to_excel(f'{path_to_end_folder_calculate_data}/Проверка вычисления {current_time}.xlsx', index=False)

        # Создание итоговой таблицы результатов подсчета

        finish_result = pd.DataFrame()

        finish_result['Наименование показателя'] = result_dct.keys()
        finish_result['Значение показателя'] = result_dct.values()
        # Проводим обработку в зависимости от значения переключателя

        # Получаем текущее время для того чтобы использовать в названии

        if mode_text == 'Yes':
            # Обрабатываем датафрейм считая текстовые данные
            count_text_df = count_text_value(finish_result)
            # сохраняем


            count_text_df.to_excel(
                f'{path_to_end_folder_calculate_data}/Подсчет текстовых значений {current_time}.xlsx')
        else:
            # сохраняем

            finish_result.to_excel(f'{path_to_end_folder_calculate_data}/Итоговые значения {current_time}.xlsx',
                                   index=False)

        if count_errors != 0:
            messagebox.showinfo('Веста Обработка таблиц и создание документов',
                                f'Обработка файлов завершена!\nОбработано файлов:  {count} из {quantity_files}\n Необработанные файлы указаны в файле {path_to_end_folder_calculate_data}/ERRORS {current_time}.txt ')
        else:
            messagebox.showinfo('Веста Обработка таблиц и создание документов',
                                f'Обработка файлов успешно завершена!\nОбработано файлов:  {count} из {quantity_files}')
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
    except FileNotFoundError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')
    # except:
    #     logging.exception('AN ERROR HAS OCCURRED')
    #     messagebox.showerror('Веста Обработка таблиц и создание документов',
    #                          'Возникла ошибка!!! Подробности ошибки в файле error.log')

if __name__ == '__main__':
    mode_text = 'Yes'
    name_file_params_calculate_data = 'data\Извлечение данных\Анкеты мониторинг профориентации\Параметры для подсчета анкет.xlsx'
    names_files_calculate_data = ['data/Извлечение данных\Анкеты мониторинг профориентации\Усть-Кяхтинская СОШ.xlsx',
                                  'data/Извлечение данных\Анкеты мониторинг профориентации\МБОУ Ацульская СОШ.xlsx']
    # names_files_calculate_data = ''
    path_to_end_folder_calculate_data = 'data'


    extract_data_from_hard_xlsx(mode_text, name_file_params_calculate_data, names_files_calculate_data,
                                path_to_end_folder_calculate_data)