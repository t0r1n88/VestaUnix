# -*- coding: utf-8 -*-
"""
Функция для подсчета текущего возраста и разбиения по возрастным категориям
"""
import pandas as pd
from tkinter import messagebox
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import time
import sys
import locale
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging

logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)



def extract_number_month(cell):
    """
    Функция для извлечения номера месяца
    """
    return cell.month


def extract_name_month(cell):
    """
    Функция для извлечения названия месяца
    Взято отсюда https://ru.stackoverflow.com/questions/1045154/Вывод-русских-символов-из-pd-timestamp-month-name
    """
    # return cell.month_name(locale='Russian')
    return cell.month_name()


def extract_year(cell):
    """
    Функция для извлечения года рождения
    """
    return cell.year


def calculate_age(born, raw_selected_date):
    """
    Функция для расчета текущего возраста взято с https://stackoverflow.com/questions/2217488/age-from-birthdate-in-python/9754466#9754466
    :param born: дата рождения
    :return: возраст
    """

    try:
        # today = date.today()
        selected_date = pd.to_datetime(raw_selected_date, dayfirst=True)
        # return today.year - born.year - ((today.month, today.day) < (born.month, born.day))
        return selected_date.year - born.year - ((selected_date.month, selected_date.day) < (born.month, born.day))

    except ValueError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Введена некорректная дата относительно которой нужно провести обработку\nПример корректной даты 01.09.2022')
        logging.exception('AN ERROR HAS OCCURRED')
        quit()



def proccessing_date(raw_selected_date, name_column, name_file_data_date, path_to_end_folder_date):
    """
   Функция для разбиения по категориям 1-ПК 1-ПО СПО-1, подсчета текущего возраста и выделения месяца,года
    :param raw_selected_date: дата на момент которой нужно подсчитать текущий возраст в формате DD.MM.YYYY
    :param name_column: название колонки с датами рождения
    :param name_file_data_date: путь к файлу Excel с данными
    :param path_to_end_folder_date: папка куда будет сохранен итоговый файл
    :return: файл Excel  содержащий исходный файл с добавленными колонками категорий и т.п.
    """

    try:

        # Считываем файл
        df = pd.read_excel(name_file_data_date)
        # Конвертируем его в формат даты
        # В случае ошибок заменяем значение NaN
        df[name_column] = pd.to_datetime(df[name_column], dayfirst=True, errors='coerce')

        # Создаем файл excel
        wb = openpyxl.Workbook()
        # Создаем листы
        # Переименовываем лист чтобы в итоговом файле не было пустого листа
        ren_sheet = wb['Sheet']
        ren_sheet.title = 'Итоговая таблица'

        # wb.create_sheet(title='Итоговая таблица', index=0)
        wb.create_sheet(title='Свод по возрастам', index=1)
        wb.create_sheet(title='Свод по месяцам', index=2)
        wb.create_sheet(title='Свод по годам', index=3)
        wb.create_sheet(title='Свод по 1-ПК', index=4)
        wb.create_sheet(title='Свод по 1-ПО', index=5)
        wb.create_sheet(title='Свод по СПО-1', index=6)
        wb.create_sheet(title='Свод по категориям Росстата', index=7)

        # Подсчитываем текущий возраст
        df['Текущий возраст'] = df[name_column].apply(lambda x:calculate_age(x, raw_selected_date))

        # Получаем номер месяца
        df['Порядковый номер месяца рождения'] = df[name_column].apply(extract_number_month)

        # Получаем название месяца
        df['Название месяца рождения'] = df[name_column].apply(extract_name_month)

        # Получаем год рождения
        df['Год рождения'] = df[name_column].apply(extract_year)

        # Присваиваем категорию по 1-ПК
        df['1-ПК Категория'] = pd.cut(df['Текущий возраст'], [0, 24, 29, 34, 39, 44, 49, 54, 59, 64, 101, 10000],
                                      labels=['моложе 25 лет', '25-29', '30-34', '35-39',
                                              '40-44', '45-49', '50-54', '55-59', '60-64',
                                              '65 и более',
                                              'Возраст  больше 101'])
        # Приводим к строковому виду, иначе не запишется на лист
        df['1-ПК Категория'] = df['1-ПК Категория'].astype(str)
        df['1-ПК Категория'] = df['1-ПК Категория'].replace('nan','Ошибочное значение!!!')

        # Присваиваем категорию по 1-ПО
        df['1-ПО Категория'] = pd.cut(df['Текущий возраст'],
                                      [0, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25,
                                       26, 27, 28,
                                       29, 34, 39, 44, 49, 54, 59, 64, 101],
                                      labels=['моложе 14 лет', '14 лет', '15 лет',
                                              '16 лет',
                                              '17 лет', '18 лет', '19 лет', '20 лет',
                                              '21 год', '22 года',
                                              '23 года', '24 года', '25 лет',
                                              '26 лет', '27 лет', '28 лет', '29 лет',
                                              '30-34 лет',
                                              '35-39 лет', '40-44 лет', '45-49 лет',
                                              '50-54 лет',
                                              '55-59 лет',
                                              '60-64 лет',
                                              '65 лет и старше'])
        # Приводим к строковому виду, иначе не запишется на лист
        df['1-ПО Категория'] = df['1-ПО Категория'].astype(str)
        df['1-ПО Категория'] = df['1-ПО Категория'].replace('nan', 'Ошибочное значение!!!')

        # Присваиваем категорию по 1-СПО
        df['СПО-1 Категория'] = pd.cut(df['Текущий возраст'],
                                       [0, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 34,
                                        39,
                                        101],
                                       labels=['моложе 13 лет', '13 лет', '14 лет', '15 лет', '16 лет', '17 лет',
                                               '18 лет',
                                               '19 лет', '20 лет'
                                           , '21 год', '22 года', '23 года', '24 года', '25 лет', '26 лет', '27 лет',
                                               '28 лет',
                                               '29 лет',
                                               '30-34 лет', '35-39 лет', '40 лет и старше'])
        ## Приводим к строковому виду, иначе не запишется на лист
        df['СПО-1 Категория'] = df['СПО-1 Категория'].astype(str)
        df['СПО-1 Категория'] = df['СПО-1 Категория'].replace('nan', 'Ошибочное значение!!!')

        # Присваиваем категорию по Росстату
        df['Росстат Категория'] = pd.cut(df['Текущий возраст'],
                                         [0, 4, 9, 14, 19, 24, 29, 34, 39, 44, 49, 54, 59, 64, 69, 200],
                                         labels=['0-4', '5-9', '10-14', '15-19', '20-24', '25-29', '30-34',
                                                 '35-39', '40-44', '45-49', '50-54', '55-59', '60-64', '65-69',
                                                 '70 лет и старше'])
        ## Приводим к строковому виду, иначе не запишется на лист
        df['Росстат Категория'] = df['Росстат Категория'].astype(str)
        df['Росстат Категория'] = df['Росстат Категория'].replace('nan', 'Ошибочное значение!!!')

        # Заполняем пустые строки
        df.fillna('Ошибочное значение!!!', inplace=True)

        # заполняем сводные таблицы
        # Сводная по возрастам

        df_svod_by_age = df.groupby(['Текущий возраст']).agg({name_column: 'count'})
        df_svod_by_age.columns = ['Количество']

        for r in dataframe_to_rows(df_svod_by_age, index=True, header=True):
            wb['Свод по возрастам'].append(r)

        # Сводная по месяцам
        df_svod_by_month = df.groupby(['Название месяца рождения']).agg({name_column: 'count'})
        df_svod_by_month.columns = ['Количество']

        # Сортируем индекс чтобы месяцы шли в хоронологическом порядке
        # Взял отсюда https://stackoverflow.com/questions/40816144/pandas-series-sort-by-month-index
        df_svod_by_month.index = pd.CategoricalIndex(df_svod_by_month.index,
                                                     categories=['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                                                                 'Июль',
                                                                 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'],
                                                     ordered=True)
        df_svod_by_month.sort_index(inplace=True)

        for r in dataframe_to_rows(df_svod_by_month, index=True, header=True):
            wb['Свод по месяцам'].append(r)

        # Сводная по годам
        df_svod_by_year = df.groupby(['Год рождения']).agg({name_column: 'count'})
        df_svod_by_year.columns = ['Количество']

        for r in dataframe_to_rows(df_svod_by_year, index=True, header=True):
            wb['Свод по годам'].append(r)

        # Сводная по 1-ПК
        df_svod_by_1PK = df.groupby(['1-ПК Категория']).agg({name_column: 'count'})
        df_svod_by_1PK.columns = ['Количество']

        for r in dataframe_to_rows(df_svod_by_1PK, index=True, header=True):
            wb['Свод по 1-ПК'].append(r)

        # Сводная по 1-ПО
        df_svod_by_1PO = df.groupby(['1-ПО Категория']).agg({name_column: 'count'})
        df_svod_by_1PO.columns = ['Количество']

        for r in dataframe_to_rows(df_svod_by_1PO, index=True, header=True):
            wb['Свод по 1-ПО'].append(r)

        # Сводная по СПО-1
        df_svod_by_SPO1 = df.groupby(['СПО-1 Категория']).agg({name_column: 'count'})
        df_svod_by_SPO1.columns = ['Количество']

        for r in dataframe_to_rows(df_svod_by_SPO1, index=True, header=True):
            wb['Свод по СПО-1'].append(r)

        # Сводная по Росстату
        df_svod_by_Ros = df.groupby(['Росстат Категория']).agg({name_column: 'count'})
        df_svod_by_Ros.columns = ['Количество']

        # Сортируем индекс
        df_svod_by_Ros.index = pd.CategoricalIndex(df_svod_by_Ros.index,
                                                   categories=['0-4', '5-9', '10-14', '15-19', '20-24', '25-29',
                                                               '30-34',
                                                               '35-39', '40-44', '45-49', '50-54', '55-59', '60-64',
                                                               '65-69',
                                                               '70 лет и старше', 'nan'],
                                                   ordered=True)
        df_svod_by_Ros.sort_index(inplace=True)

        for r in dataframe_to_rows(df_svod_by_Ros, index=True, header=True):
            wb['Свод по категориям Росстата'].append(r)

        for r in dataframe_to_rows(df, index=False, header=True):
            wb['Итоговая таблица'].append(r)

        # сохраняем по ширине колонок
        for column in wb['Итоговая таблица'].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb['Итоговая таблица'].column_dimensions[column_name].width = adjusted_width

        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        wb.save(f'{path_to_end_folder_date}/Результат обработки колонки {name_column} от {current_time}.xlsx')
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файл с данными и папку куда будет генерироваться файл')
        logging.exception('AN ERROR HAS OCCURRED')
    except KeyError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'В таблице нет такой колонки!\nПроверьте написание названия колонки')
        logging.exception('AN ERROR HAS OCCURRED')
    except FileNotFoundError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')

    # except:
    #     logging.exception('AN ERROR HAS OCCURRED')
    #     messagebox.showerror('Веста Обработка таблиц и создание документов',
    #                          'Возникла ошибка!!! Подробности ошибки в файле error.log')
    else:
        messagebox.showinfo('Веста Обработка таблиц и создание документов', 'Данные успешно обработаны')

if __name__ == '__main__':
    raw_selected_date_main = '01.10.2023'
    name_column_main = 'Дата рождения'
    name_file_data_date_main = 'data/Обработка дат/Сгенерированный массив данных для дат.xlsx'
    path_to_end_folder_date_main = 'data'
    proccessing_date(raw_selected_date_main, name_column_main, name_file_data_date_main, path_to_end_folder_date_main)
    print('Lindy Booth')
