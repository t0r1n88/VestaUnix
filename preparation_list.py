"""
Скрипт для подготовки списка
Очистка некорректных данных, удаление лишних пробелов
"""
from support_functions import write_df_to_excel # функция для записи в файл Excel с автоподбором ширины
import time
import gc
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import datetime
import re
from tkinter import messagebox
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None

class ExceedingQuantity(Exception):
    """
    Исключение для случаев когда числа уникальных значений больше 255
    """
    pass


def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        if cell is np.nan:
            return 'Не заполнено'
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except ValueError:
        return f'Неправильное значение - {cell}'
    except TypeError:
        return f'Неправильное значение - {cell}'


def capitalize_fio(value:str)->str:
    """
    Функция для применения capitalize к значениям состоящим из несколько слов разделенных пробелами
    value: значение ячейки
    """
    value = str(value)
    if value == 'Не заполнено':
        return value
    temp_lst = value.split(' ') # создаем список по пробелу
    temp_lst = list(map(str.capitalize,temp_lst))  # обрабатываем
    return ' '.join(temp_lst) #соединяем в строку


def find_english_letter(value):
    """
    Функция для поиска английских букв в ФИО
    :param value: строка ФИО
    :return:
    """
    result = re.findall(r'[a-zA-Z]',value)
    if result:
        english_let = ';'.join(result)
        return f'Обнаружены символы латиницы: {english_let} в слове {value}'
    else:
        return value

def prepare_fio_text_columns(df:pd.DataFrame,lst_columns:list)->pd.DataFrame:
    """
    Функция для очистки текстовых колонок c данными ФИО
    df: датафрейм для обработки
    lst_columns: список колонок которые нужно обработать
    """
    prepared_columns_lst = [] # список для колонок содержащих слова Фамилия,Имя,Отчество, ФИО
    for fio_column in lst_columns:
        for name_column in df.columns:
            if fio_column in name_column.lower():
                prepared_columns_lst.append(name_column)
    if len(prepared_columns_lst) == 0: # проверка на случай не найденных значений
        return df

    df[prepared_columns_lst] = df[prepared_columns_lst].fillna('Не заполнено')
    df[prepared_columns_lst] = df[prepared_columns_lst].astype(str)
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(lambda x: x.strip() if isinstance(x, str) else x)  # применяем strip, чтобы все данные корректно вставлялись
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(lambda x:' '.join(x.split())) # убираем лишние пробелы между словами
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(capitalize_fio)  # делаем заглавными первые буквы слов а остальыне строчными
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(find_english_letter)  # делаем заглавными первые буквы слов а остальыне строчными

    return df

def prepare_date_column(df:pd.DataFrame,lst_columns:list)->pd.DataFrame:
    """
    Функция для обработки колонок с датами
    df: датафрейм для обработки
    lst_columns: список колонок которые нужно обработать
    """
    prepared_columns_lst = [] # список для колонок содержащих слово дата
    for date_column in lst_columns:
        for name_column in df.columns:
            if date_column in name_column.lower():
                prepared_columns_lst.append(name_column)
    if len(prepared_columns_lst) == 0: # проверка на случай не найденных значений
        return df
    df[prepared_columns_lst] = df[prepared_columns_lst].fillna('Не заполнено')
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(lambda x:pd.to_datetime(x,errors='ignore',dayfirst=True)) # приводим к типу дата
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(create_doc_convert_date)  # приводим к виду ДД.ММ.ГГГГ
    return df

def prepare_snils(df:pd.DataFrame,snils:str)->pd.DataFrame:
    """
    Функция для обработки колонок со снилс
    df: датафрейм для обработки
    snils: название снилс
    """

    prepared_columns_lst = []  # список для колонок содержащих слово снилс
    for name_column in df.columns:
        if snils in name_column.lower():
            prepared_columns_lst.append(name_column)

    if len(prepared_columns_lst) == 0: # проверка на случай не найденных значений
        return df

    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(check_snils)

    return df

def prepare_snils_copp(df:pd.DataFrame,snils:str)->pd.DataFrame:
    """
    Функция для обработки колонок со снилс
    df: датафрейм для обработки
    snils: название снилс
    """
    if snils not in df.columns:
        messagebox.showerror('','Не найдена колонка СНИЛС!!!')

    df['СНИЛС'] =df['СНИЛС'].apply(check_snils)
    return df



def check_snils(snils):
    """
    Функция для приведения значений снилс в вид ХХХ-ХХХ-ХХХ ХХ
    """
    if snils is np.nan:
        return 'Не заполнено'
    snils = str(snils)
    result = re.findall(r'\d', snils) # ищем цифры
    if len(result) == 11:
        first_group = ''.join(result[:3])
        second_group = ''.join(result[3:6])
        third_group = ''.join(result[6:9])
        four_group = ''.join(result[9:11])

        out_snils = f'{first_group}-{second_group}-{third_group} {four_group}'
        return out_snils
    else:
        return f'Неправильное значение!В СНИЛС должно быть 11 цифр - {snils} -{len(snils)} цифр'

def prepare_inn_column(df:pd.DataFrame,lst_columns:list)->pd.DataFrame:
    """
    Функция для обработки колонок со снилс
    df: датафрейм для обработки
    lst_columns: список колонок с ИНН
    """

    prepared_columns_lst = [] # список для колонок содержащих слово дата
    for inn_column in lst_columns:
        for name_column in df.columns:
            if inn_column in name_column.lower():
                prepared_columns_lst.append(name_column)
    if len(prepared_columns_lst) == 0: # проверка на случай не найденных значений
        return df

    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(check_inn) # обрабатываем инн
    return df


def check_inn(inn):
    """
    Функция для приведения значений снилс в вид 12 цифр
    """
    if inn is np.nan:
        return 'Не заполнено'
    inn = str(inn)
    result = re.findall(r'\d', inn) # ищем цифры
    if len(result) == 12:
        return ''.join(result)
    else:
        return f'Неправильное значение ИНН (ИНН физлица состоит из 12 цифр)- {inn} -{len(inn)} цифр'

def prepare_passport_column(df:pd.DataFrame)->pd.DataFrame:
    """
    Функция для обработки колонок серия и номер паспорта
    df: датафрейм для обработки
    series_passport: значение для поиска колонкок с содержащей серию паспорта
    number_passport: значение для поиска колонкок с содержащей серию паспорта
    code_passport: значение для поиска колонкок с содержащей код подразделения

    """
    prepared_columns_series_lst = [] # список для колонок содержащих слова серия паспорт
    prepared_columns_number_lst = [] # список для колонок содержащих слова номер паспорт
    prepared_columns_code_lst = [] # список для колонок содержащих слова код подразд
    pattern_series = re.compile(r"(?=.*серия)(?=.*паспорт)") # паттерн для серии паспорта
    pattern_number = re.compile(r"(?=.*номер)(?=.*паспорт)") # паттерн для номера паспорта
    pattern_code = re.compile(r"(?=.*код)(?=.*подразд)") # паттерн для кода подразделения
    for name_column in df.columns:
        result_series = re.search(pattern_series,name_column.lower()) # ищем по паттерну серию
        if result_series:
            prepared_columns_series_lst.append(name_column)
        result_number = re.search(pattern_number,name_column.lower()) # ищем по паттерну номер
        if result_number:
            prepared_columns_number_lst.append(name_column)
        result_code =   re.search(pattern_code,name_column.lower()) # ищем по паттерну код подразделения
        if result_code:
            prepared_columns_code_lst.append(name_column)


    if len(prepared_columns_series_lst) != 0:
        df[prepared_columns_series_lst] = df[prepared_columns_series_lst].applymap(check_series_passport)  # обрабатываем серию паспорта

    if len(prepared_columns_number_lst) != 0:
        df[prepared_columns_number_lst] = df[prepared_columns_number_lst].applymap(check_number_passport)  # обрабатываем номер паспорта

    if len(prepared_columns_code_lst) != 0:
        df[prepared_columns_code_lst] = df[prepared_columns_code_lst].applymap(check_code_passport)  # обрабатываем код подразделения

    return df

def check_series_passport(series:str)->str:
    """
    Функция для проверки серии паспорта, должно быть 4 цифры
    """
    if series is np.nan:
        return 'Не заполнено'
    series = str(series)
    result = re.findall(r'\d', series) # ищем цифры
    if len(result) == 4:
        return ''.join(result)
    else:
        return f'Неправильное значение серии паспорта (должно быть 4 цифры) - {series}'

def check_number_passport(number:str)->str:
    """
    Функция для проверки номера паспорта, должно быть 6 цифр
    """
    if number is np.nan:
        return 'Не заполнено'
    number = str(number)
    result = re.findall(r'\d', number) # ищем цифры
    if len(result) == 6:
        return ''.join(result)
    else:
        return f'Неправильное значение номера паспорта(должно быть 6 цифр) - {number}'

def check_code_passport(code:str)->str:
    """
    Функция для проверки номера паспорта, должно быть 6 цифр
    """
    if code is np.nan:
        return 'Не заполнено'
    code = str(code)
    result = re.findall(r'\d', code) # ищем цифры
    if len(result) == 6:
        first_group = ''.join(result[:3])
        second_group = ''.join(result[3:6])
        return f'{first_group}-{second_group}'
    else:
        return f'Неправильное значение кода подразделения(должно быть 6 цифр в формате XXX-XXX) - {code}'

def prepare_phone_columns(df:pd.DataFrame,phone_text:str) ->pd.DataFrame:
    """
    Функция для очистки номеров телефонов от пробельных символов и букв
    """
    # pattern = r'[a-zA-Zа-яА-Я\s.]'
    pattern = r'\D' # удаляем  все кроме цифр
    prepared_phone_columns = [] # лист для колонок с телефонами
    # собираем названия колонок содержащих слово телефон
    for name_column in df.columns:
        if phone_text in name_column.lower():
            prepared_phone_columns.append(name_column)

    if len(prepared_phone_columns) == 0:
        return df

    df[prepared_phone_columns] = df[prepared_phone_columns].applymap(lambda x:check_phone_number(x,pattern))
    return df

def check_phone_number(phone:str,pattern:str)->str:
    """
    Функция для очистки значения номера телефона от пробельных символов,букв и точки
    """
    if phone is np.nan:
        return 'Не заполнено'
    phone = str(phone)
    clean_phone = re.sub(pattern,'',phone)
    return clean_phone


def prepare_email_columns(df:pd.DataFrame,second_option:str)->pd.DataFrame:
    """
    Функция для обработки колонок серия и номер паспорта
    df: датафрейм для обработки
    second_option: значение для поиска колонкок с содержащей слово e-mail
    """
    prepared_columns_email_set = set() # множество для колонок содержащих слова электрон почта e-mail
    # это нужно для обработки случаем когда колонка называется Электронная почта(email)
    pattern_first_option = re.compile(r"(?=.*электрон)(?=.*почта)") # паттерн для слов электрон почта
    for name_column in df.columns:
        result_first_option = re.search(pattern_first_option,name_column.lower()) # ищем по паттерну электрон почта
        if result_first_option:
            prepared_columns_email_set.add(name_column)
        if second_option in name_column:
            prepared_columns_email_set.add(name_column)
    prepared_columns_email_lst = list(prepared_columns_email_set) # превращаем в список

    if len(prepared_columns_email_lst) == 0:
        return df
    df[prepared_columns_email_lst] = df[prepared_columns_email_lst].fillna('Не заполнено')
    df[prepared_columns_email_lst] = df[prepared_columns_email_lst].applymap(lambda x:re.sub(r'\s','',x) if x !='Не заполнено' else x)

    return df


def prepare_list(file_data:str,path_end_folder:str,checkbox_dupl:str):
    """
    file_data : путь к файлу который нужно преобразовать
    path_end_folder :  путь к конечной папке
    checkbox_dupl: Проверять на дубликаты или нет. Yes or No
    """
    try:
        df = pd.read_excel(file_data,dtype=str) # считываем датафрейм
        df.columns = list(map(str,list(df.columns))) # делаем названия колонок строкововыми
        # очищаем все строковые значения от пробелов в начале и конце
        df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        # обрабатываем колонки с фио
        part_fio_columns = ['фамилия','имя','отчество','фио'] # колонки с типичными названиями
        df = prepare_fio_text_columns(df,part_fio_columns) # очищаем колонки с фио

        # обрабатываем колонки содержащими слово дата
        part_date_columns = ['дата']
        df = prepare_date_column(df,part_date_columns)

        # обрабатываем колонки со снилс
        snils = 'снилс'
        df = prepare_snils(df, snils)

        # обрабатываем колонки с ИНН
        part_inn_columns = ['инн']
        df = prepare_inn_column(df,part_inn_columns)

        # обрабатываем колонки данные паспорта
        df = prepare_passport_column(df)

        # обрабатываем  колонки с номера телефонов
        phone = 'телефон'
        df = prepare_phone_columns(df, phone)

        # очищаем email от пробельных символов
        second_option = 'e-mail' # слова электрон и почта используются внутри функции
        df = prepare_email_columns(df,second_option)
        # получаем время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        if checkbox_dupl == 'Yes':
            """
            Создаем список дубликатов
            """
            lst_name_columns = list(df.columns)  # получаем список колонок
            used_name_sheet = []  # список для хранения значений которые уже были использованы
            if len(lst_name_columns) >= 253:  # проверяем количество колонок которые могут созданы
                raise ExceedingQuantity
            #
            wb = openpyxl.Workbook(write_only=True)  # создаем файл
            for idx, value in enumerate(lst_name_columns):
                temp_df = df[df[value].duplicated(keep=False)]  # получаем дубликаты
                if temp_df.shape[0] == 0:
                    continue

                short_value = value[:20]  # получаем обрезанное значение
                short_value = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_value)

                if short_value in used_name_sheet:
                    short_value = f'{short_value}_{idx}'  # добавляем окончание
                wb.create_sheet(short_value, index=idx)  # создаем лист
                used_name_sheet.append(short_value)

                temp_df = temp_df.sort_values(by=value)
                #     # Добавляем +2 к индексу чтобы отобразить точную строку
                temp_df.insert(0, '№ строки дубликата ', list(map(lambda x: x + 2, list(temp_df.index))))

                for row in dataframe_to_rows(temp_df, index=False, header=True):
                    wb[short_value].append(row)

            wb.save(f'{path_end_folder}/Дубликаты в каждой колонке {current_time}.xlsx')
            # очищаем
            wb.close()
            del wb
            gc.collect()

        # сохраняем

        dct_df = {'Лист1': df}
        write_index = False
        wb_main = write_df_to_excel(dct_df, write_index)
        name_file = file_data.split('.xlsx')[0]  # получаем путь без расширения
        name_file = name_file.split('/')[-1]
        wb_main.save(f'{path_end_folder}/Обработанный {name_file} {current_time}.xlsx')
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
        logging.exception('AN ERROR HAS OCCURRED')

    except ValueError as e:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Ошибка при обработке значения {e.args}')
        logging.exception('AN ERROR HAS OCCURRED')

    # except FileNotFoundError:
    #     messagebox.showerror('Веста Обработка таблиц и создание документов',
    #                          f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
    #                          f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')
    else:
        messagebox.showinfo('Веста Обработка таблиц и создание документов', 'Данные успешно обработаны')


if __name__ == '__main__':
    # file_data_main = 'data/Обработка списка/Список студентов военкомат.xlsx'
    file_data_main = 'data/Обработка списка/Список студентов военкомат.xlsx'
    path_end_main = 'data'
    checkbox_main = 'Yes'
    start_time = time.time()
    prepare_list(file_data_main,path_end_main,checkbox_main)
    end_time = time.time()
    execution_time = end_time - start_time
    print(f"Время выполнения: {execution_time} секунд")
    print('Lindy Booth')

