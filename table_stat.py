"""
Функции для подсчета категориальных (количество значений в колонке) и количественных статистик таблицы
"""
import pandas as pd
from tkinter import messagebox
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import warnings
from collections import Counter
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging
import re
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)

def counting_by_category(name_file_data_groupby,path_to_end_folder_groupby):
    """
    Функция для подсчета всех колонок таблицы по категориям
    :param name_file_data_groupby: путь к файлу
    :param path_to_end_folder_groupby:  путь к итоговой папке
    """
    try:
        df = pd.read_excel(name_file_data_groupby)
        df.columns = list(map(str, list(df.columns)))

        # Создаем файл excel
        wb = openpyxl.Workbook()

        # Проверяем наличие возможных дубликатов ,котороые могут получиться если обрезать по 30 символов
        lst_length_column = [column[:30] for column in df.columns]
        check_dupl_length = [k for k, v in Counter(lst_length_column).items() if v > 1]

        # проверяем наличие объединенных ячеек
        check_merge = [column for column in df.columns if 'Unnamed' in column]
        # если есть хоть один Unnamed то просто заменяем названия колонок на Колонка №цифра
        if check_merge or check_dupl_length:
            df.columns = [f'Колонка №{i}' for i in range(1, df.shape[1] + 1)]
        # очищаем названия колонок от символов */\ []''
        # Создаем регулярное выражение
        pattern_symbols = re.compile(r"[/*'\[\]/\\]")
        clean_df_columns = [re.sub(pattern_symbols, '', column) for column in df.columns]
        df.columns = clean_df_columns

        # Добавляем столбец для облегчения подсчета по категориям
        df['Для подсчета'] = 1

        # Создаем листы
        for idx, name_column in enumerate(df.columns):
            # Делаем короткое название не более 30 символов
            wb.create_sheet(title=name_column[:30], index=idx)

        for idx, name_column in enumerate(df.columns):
            group_df = df.groupby([name_column]).agg({'Для подсчета': 'sum'})
            group_df.columns = ['Количество']

            # Сортируем по убыванию
            group_df.sort_values(by=['Количество'], inplace=True, ascending=False)

            for r in dataframe_to_rows(group_df, index=True, header=True):
                if len(r) != 1:
                    wb[name_column[:30]].append(r)
            wb[name_column[:30]].column_dimensions['A'].width = 50

        # генерируем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Удаляем листы
        del wb['Sheet']
        del wb['Для подсчета']
        # Сохраняем итоговый файл
        wb.save(
            f'{path_to_end_folder_groupby}/Подсчет частоты значений для всех колонок таблицы от {current_time}.xlsx')

    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файл с данными и папку куда будет генерироваться файл')

    except FileNotFoundError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except:
        logging.exception('AN ERROR HAS OCCURRED')
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             'Возникла ошибка!!! Подробности ошибки в файле error.log')
    else:
        messagebox.showinfo('Веста Обработка таблиц и создание документов', 'Данные успешно обработаны')


def counting_quantitative_stat(name_file_data_groupby,path_to_end_folder_groupby):
    """
    Функция для подсчета выбранной колонки по количественным показателям(сумма,среднее,медиана,мин,макс)
    :param name_file_data_groupby: путь к файлу
    :param path_to_end_folder_groupby: путь к итоговой папке
    :return:
    """
    try:
        df = pd.read_excel(name_file_data_groupby)
        # Делаем названия колонок строковыми
        df.columns = list(map(str, list(df.columns)))
        # Создаем файл excel
        wb = openpyxl.Workbook()

        # Проверяем наличие возможных дубликатов ,котороые могут получиться если обрезать по 30 символов
        lst_length_column = [column[:30] for column in df.columns]
        check_dupl_length = [k for k, v in Counter(lst_length_column).items() if v > 1]

        # проверяем наличие объединенных ячеек
        check_merge = [column for column in df.columns if 'Unnamed' in column]
        # если есть хоть один Unnamed или дубликат то просто заменяем названия колонок на Колонка №цифра
        if check_merge or check_dupl_length:
            df.columns = [f'Колонка №{i}' for i in range(1, df.shape[1] + 1)]

        # очищаем названия колонок от символов */\ []''
        # Создаем регулярное выражение
        pattern_symbols = re.compile(r"[/*'\[\]/\\]")
        clean_df_columns = [re.sub(pattern_symbols, '', column) for column in df.columns]
        df.columns = clean_df_columns

        # Добавляем столбец для облегчения подсчета по категориям
        df['Итого'] = 1

        # Создаем листы
        for idx, name_column in enumerate(df.columns):
            # Делаем короткое название не более 30 символов
            wb.create_sheet(title=name_column[:30], index=idx)

        for idx, name_column in enumerate(df.columns):
            group_df = df[name_column].describe().to_frame()
            if group_df.shape[0] == 8:
                # подсчитаем сумму
                all_sum = df[name_column].sum()
                dct_row = {name_column: all_sum}
                row = pd.DataFrame(data=dct_row, index=['Сумма'])
                # Добавим в датафрейм
                group_df = pd.concat([group_df, row], axis=0)

                # Обновим названия индексов
                group_df.index = ['Количество значений', 'Среднее', 'Стандартное отклонение', 'Минимальное значение',
                                  '25%(Первый квартиль)', 'Медиана', '75%(Третий квартиль)', 'Максимальное значение',
                                  'Сумма']

            elif group_df.shape[0] == 4:
                group_df.index = ['Количество значений', 'Количество уникальных значений', 'Самое частое значение',
                                  'Количество повторений самого частого значения', ]
            for r in dataframe_to_rows(group_df, index=True, header=True):
                if len(r) != 1:
                    wb[name_column[:30]].append(r)
            wb[name_column[:30]].column_dimensions['A'].width = 50

        # генерируем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)
        # Удаляем лист
        del wb['Sheet']
        del wb['Итого']
        # Сохраняем итоговый файл
        wb.save(
            f'{path_to_end_folder_groupby}/Подсчет базовых статистик для всех колонок таблицы от {current_time}.xlsx')

    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файл с данными и папку куда будет генерироваться файл')
        logging.exception('AN ERROR HAS OCCURRED')

    except FileNotFoundError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')
    except:
        logging.exception('AN ERROR HAS OCCURRED')
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             'Возникла ошибка!!! Подробности ошибки в файле error.log')

    else:
        messagebox.showinfo('Веста Обработка таблиц и создание документов', 'Данные успешно обработаны')

if __name__ == '__main__':
    name_file_data_groupby_main = 'data\Подсчет данных\Пример таблицы для подсчета.xlsx'
    path_to_end_folder_groupby_main = 'data'
    counting_by_category(name_file_data_groupby_main, path_to_end_folder_groupby_main)
    counting_quantitative_stat(name_file_data_groupby_main, path_to_end_folder_groupby_main)
    print('Lindy Booth')