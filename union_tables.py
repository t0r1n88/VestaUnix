"""
Функция для объединения нескольких таблиц в одну
"""
from support_functions import write_df_to_excel # импорт функции по записи в файл с автошириной колонок
import pandas as pd
import os
from tkinter import messagebox
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import time
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
from jinja2 import exceptions
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)

class NotParamsHarvest(BaseException):
    """
    Проверяет наличие файла с параметрами объединения файлов
    """
    pass


def union_tables(checkbox_harvest: int,merger_entry_skip_rows: int, file_standard_merger:str,dir_name: str,path_to_end_folder_merger:str,params_harvest:str):
    """
    Функция для слияния таблиц с одинаковой структурой в одну большую таблицу
    :param checkbox_harvest: Чекбокс отвечающий за вариант слияния. Допустимые значения 0 - слияние по названию листов,
    1 - слияние по порядковому номеру листов,2 - сложное слияние листов с разным размеров заголовков
    :param merger_entry_skip_rows: Количество строк в таблице которое занимает заголовок
    :param file_standard_merger:Путь к файлу эталону на основе которого будет создаваться общая таблица
    :param dir_name: Папка где лежат файлы которые нужно объединить
    :param path_to_end_folder_merger: Папка куда будут сохранены результаты
    :param params_harvest: Путь к файлу Excel  в котором указаны названия листов и размер заголовков на каждом листе.
     Используется при слиянии по варианту 2 (Сложное)
    :return:Сохраняет 2 файла: Общий файл с данными из всех файлов, Файл с ошибками в которых указаны те файлы Excel которые отличаются от эталонного
    """

    # Получаем значения из полей ввода и проверяем их на тип
    try:
        if checkbox_harvest != 2:
            skip_rows = int(merger_entry_skip_rows)
    except ValueError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             'Введите целое число в поле для ввода количества пропускаемых строк!!!')
    else:
        # Оборачиваем в try
        try:
            # Создаем датафрейм куда будем сохранять ошибочные файлы
            err_df = pd.DataFrame(columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'])

            name_file_standard_merger = file_standard_merger.split('/')[-1]  # получаем имя файла

            standard_wb = load_workbook(filename=file_standard_merger)  # Загружаем эталонный файл

            standard_sheets = sorted(
                standard_wb.sheetnames)  # отсортрованный список листов по которому будет вестись сравнение
            set_standard_sheets = set(standard_sheets)  # создаем множество из листов эталонного файла
            standard_size_sheets = len(standard_sheets)

            "Удаляем пустые и строки с заливкой которые могут тянуться вниз и из этого данные из других файлов начина" \
            "ются с тысячных строк"
            for sheet in standard_wb.sheetnames:
                del_cols_df = pd.read_excel(file_standard_merger,
                                            sheet_name=sheet)  # загружаем датафрейм чтобы узнать сколько есть заполненны строк

                temp_sheet_max_row = standard_wb[sheet].max_row  # получаем последнюю строку
                standard_wb[sheet].delete_rows(del_cols_df.shape[0] + 2, temp_sheet_max_row)  # удаляем все лишнее

            dct_df = dict()  # создаем словарь в котором будем хранить да

            for sheet in standard_wb.sheetnames:  # Добавляем в словарь датафреймы
                temp_df = pd.read_excel(file_standard_merger, sheet_name=sheet, dtype=str)
                dct_df[sheet] = temp_df

            if checkbox_harvest == 0:  # Вариант объединения по названию листов
                for dirpath, dirnames, filenames in os.walk(dir_name):
                    for filename in filenames:
                        if (filename.endswith('.xlsx') and not filename.startswith(
                                '~$')) and filename != name_file_standard_merger:  # не обрабатываем эталонный файл
                            # Получаем название файла без расширения
                            name_file = filename.split('.xlsx')[0]
                            print(name_file)
                            temb_wb = load_workbook(
                                filename=f'{dirpath}/{filename}')  # загружаем файл, для проверки листов
                            """
                            Проверяем наличие листов из эталонного файла в проверяемом файле, если они есть то начинаем 
                            дальнейшую проверку
                            """
                            if set_standard_sheets.issubset(set(temb_wb.sheetnames)):
                                count_errors = 0
                                # проверяем наличие листов указанных в файле параметров
                                for name_sheet, df in dct_df.items():  # Проводим проверку на совпадение
                                    lst_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=name_sheet)
                                    if lst_df.shape[1] != df.shape[1]:
                                        # если количество колонок не совпадает то записываем как ошибку
                                        temp_error_df = pd.DataFrame(
                                            columns=['Название файла', 'Наименование листа', 'Тип ошибки',
                                                     'Описание ошибки'],
                                            data=[[name_file, name_sheet, 'Количество колонок отличается от эталонного',
                                                   f'Ожидалось {df.shape[1]} колонок, а в листе {lst_df.shape[1]}']])  # создаем временный датафрейм. потом надо подумать над словарем

                                        err_df = pd.concat([err_df, temp_error_df],
                                                           ignore_index=True)  # добавляем в датафрейм ошибок
                                        count_errors += 1

                                # если хоть одна ошибка то проверяем следующий файл
                                if count_errors != 0:
                                    continue
                                # если нет то начинаем обрабатывать листы
                                for name_sheet, df in dct_df.items():
                                    temp_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=name_sheet,
                                                            dtype=str, skiprows=skip_rows,
                                                            header=None)  # загружаем датафрейм
                                    if temp_df.shape[1] > 3:
                                        temp_df = temp_df.dropna(axis=0, thresh=2)

                                    temp_df['Номер строки'] = range(1, temp_df.shape[0] + 1)
                                    temp_df['Откуда взяты данные'] = name_file
                                    for row in dataframe_to_rows(temp_df, index=False, header=False):
                                        standard_wb[name_sheet].append(row)  # добавляем данные

                            elif len(
                                    temb_wb.sheetnames) == standard_size_sheets:  # сравниваем количество листов в файле
                                diff_name_sheets = set(temb_wb.sheetnames).difference(
                                    set(standard_sheets))  # проверяем разницу в названиях листов
                                print(diff_name_sheets)
                                if len(diff_name_sheets) != 0:  # если разница в названиях есть то записываем в ошибки и обрабатываем следующий файл
                                    temp_error_df = pd.DataFrame(
                                        columns=['Название файла', 'Наименование листа', 'Тип ошибки',
                                                 'Описание ошибки'], data=[
                                            [name_file, '', 'Названия листов отличаются от эталонных',
                                             f'Отличаются следующие названия листов {diff_name_sheets}']])  # создаем временный датафрейм. потом надо подумать над словарем

                                    err_df = pd.concat([err_df, temp_error_df],
                                                       ignore_index=True)  # добавляем в датафрейм ошибок

                                    continue

                            else:
                                temp_error_df = pd.DataFrame(
                                    columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'],
                                    data=[[name_file, '', 'Не совпадает количество или название листов в файле',
                                           f'Листы, которые есть в файле: {",".join(temb_wb.sheetnames)}']])  # создаем временный датафрейм. потом надо подумать над словарем

                                err_df = pd.concat([err_df, temp_error_df],
                                                   ignore_index=True)  # добавляем в датафрейм ошибок

                # Получаем текущую дату
                current_time = time.strftime('%H_%M_%S %d.%m.%Y')
                # сохраняем по ширине колонок
                first_sheet = standard_wb.sheetnames[0]
                for column in standard_wb[first_sheet].columns:
                    max_length = 0
                    column_name = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    standard_wb[first_sheet].column_dimensions[column_name].width = adjusted_width

                standard_wb.save(
                    f'{path_to_end_folder_merger}/Слияние по варианту А Общая таблица от {current_time}.xlsx')  # сохраняем
                err_out_wb = openpyxl.Workbook()  # создаем объект openpyxl для сохранения датафрейма
                for row in dataframe_to_rows(err_df, index=False, header=True):
                    err_out_wb['Sheet'].append(row)  # добавляем данные
                # устанавливаем размер колонок
                err_out_wb['Sheet'].column_dimensions['A'].width = 40
                err_out_wb['Sheet'].column_dimensions['B'].width = 30
                err_out_wb['Sheet'].column_dimensions['C'].width = 55
                err_out_wb['Sheet'].column_dimensions['D'].width = 100
                err_out_wb.save(f'{path_to_end_folder_merger}/Слияние по варианту А Ошибки от {current_time}.xlsx')

            elif checkbox_harvest == 1:  # Вариант объединения по порядку
                for dirpath, dirnames, filenames in os.walk(dir_name):
                    for filename in filenames:
                        if (filename.endswith('.xlsx') and not filename.startswith(
                                '~$')) and filename != name_file_standard_merger:  # не обрабатываем эталонный файл
                            # Получаем название файла без расширения
                            name_file = filename.split('.xlsx')[0]
                            print(name_file)
                            temb_wb = load_workbook(filename=f'{dirpath}/{filename}')  # загружаем файл

                            if standard_size_sheets == len(
                                    temb_wb.sheetnames):  # если количество листов одинаково то обрабатываем
                                count_errors = 0  # счетчик ошибок
                                dct_name_sheet = {}  # создаем словарь где ключ это название листа в эталонном файле а значение это название листа в обрабатываемом файле
                                for idx, data in enumerate(dct_df.items()):  # Проводим проверку на совпадение
                                    name_sheet = data[0]  # получаем название листа
                                    df = data[1]  # получаем датафрейм
                                    temp_name_sheet = temb_wb.sheetnames[idx]  #
                                    lst_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=temp_name_sheet)
                                    if lst_df.shape[1] != df.shape[1]:
                                        # если количество колонок не совпадает то записываем как ошибку
                                        temp_error_df = pd.DataFrame(
                                            columns=['Название файла', 'Наименование листа', 'Тип ошибки',
                                                     'Описание ошибки'],
                                            data=[
                                                [name_file, name_sheet, 'Количество колонок отличается от эталонного',
                                                 f'Ожидалось {df.shape[1]} колонок, а в листе {lst_df.shape[1]}']])  # создаем временный датафрейм. потом надо подумать над словарем

                                        err_df = pd.concat([err_df, temp_error_df],
                                                           ignore_index=True)  # добавляем в датафрейм ошибок
                                        count_errors += 1

                                    else:
                                        dct_name_sheet[name_sheet] = temp_name_sheet
                                # если хоть одна ошибка то проверяем следующий файл
                                if count_errors != 0:
                                    continue
                                    # если нет то начинаем обрабатывать листы
                                for name_sheet, df in dct_df.items():
                                    temp_df = pd.read_excel(f'{dirpath}/{filename}',
                                                            sheet_name=dct_name_sheet[name_sheet],
                                                            dtype=str, skiprows=skip_rows,
                                                            header=None)  # загружаем датафрейм
                                    if temp_df.shape[1] > 3:
                                        temp_df = temp_df.dropna(axis=0, thresh=2)
                                    temp_df['Номер строки'] = range(1, temp_df.shape[0] + 1)
                                    temp_df['Откуда взяты данные'] = name_file
                                    for row in dataframe_to_rows(temp_df, index=False, header=False):
                                        standard_wb[name_sheet].append(row)  # добавляем данные
                            else:
                                temp_error_df = pd.DataFrame(
                                    columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'],
                                    data=[[name_file, '', 'Не совпадает количество или название листов в файле',
                                           f'Листы, которые есть в файле: {",".join(temb_wb.sheetnames)}']])  # создаем временный датафрейм. потом надо подумать над словарем

                                err_df = pd.concat([err_df, temp_error_df],
                                                   ignore_index=True)  # добавляем в датафрейм ошибок

                # Получаем текущую дату
                current_time = time.strftime('%H_%M_%S %d.%m.%Y')
                # сохраняем по ширине колонок
                first_sheet = standard_wb.sheetnames[0]
                for column in standard_wb[first_sheet].columns:
                    max_length = 0
                    column_name = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    standard_wb[first_sheet].column_dimensions[column_name].width = adjusted_width
                standard_wb.save(
                    f'{path_to_end_folder_merger}/Слияние по варианту Б Общая таблица от {current_time}.xlsx')  # сохраняем

                err_out_wb = openpyxl.Workbook()  # создаем объект openpyxl для сохранения датафрейма
                for row in dataframe_to_rows(err_df, index=False, header=True):
                    err_out_wb['Sheet'].append(row)  # добавляем данные
                # устанавливаем размер колонок
                err_out_wb['Sheet'].column_dimensions['A'].width = 40
                err_out_wb['Sheet'].column_dimensions['B'].width = 30
                err_out_wb['Sheet'].column_dimensions['C'].width = 55
                err_out_wb['Sheet'].column_dimensions['D'].width = 100
                err_out_wb.save(f'{path_to_end_folder_merger}/Слияние по варианту Б Ошибки от {current_time}.xlsx')

            # Если выбран управляемый сбор данных
            elif checkbox_harvest == 2:
                df_params = pd.read_excel(params_harvest, header=None)  # загружаем параметры
                df_params[0] = df_params[0].astype(
                    str)  # делаем данные строковыми чтобы корректно работало обращение по названию листов

                tmp_name_sheets = df_params[0].tolist()  # создаем списки чтобы потом из них сделать словарь
                tmp_skip_rows = df_params[1].tolist()
                dct_manage_harvest = dict(zip(tmp_name_sheets,
                                              tmp_skip_rows))  # создаем словарь где ключ это название листа а значение это сколько строк нужно пропустить
                set_params_sheets = set(
                    dct_manage_harvest.keys())  # создаем множество из ключей(листов) которые нужно обработать
                if not set_params_sheets.issubset(
                        set_standard_sheets):  # проверяем совпадение названий в эталонном файле и в файле параметров
                    diff_value = set(dct_manage_harvest.keys()).difference(set(standard_sheets))  # получаем разницу

                    messagebox.showerror('',
                                         f'Не совпадают следующие названия листов в файле параметров и в эталонном файле\n'
                                         f'{diff_value}!')
                # начинаем обработку
                for dirpath, dirnames, filenames in os.walk(dir_name):
                    for filename in filenames:
                        if (filename.endswith('.xlsx') and not filename.startswith(
                                '~$')) and filename != name_file_standard_merger:  # не обрабатываем эталонный файл
                            # Получаем название файла без расширения
                            name_file = filename.split('.xlsx')[0]
                            print(name_file)
                            temb_wb = load_workbook(filename=f'{dirpath}/{filename}')  # загружаем файл
                            if set_params_sheets.issubset(set(temb_wb.sheetnames)):
                                count_errors = 0
                                # проверяем наличие листов указанных в файле параметров
                                for name_sheet, skip_r in dct_manage_harvest.items():  # Проводим проверку на совпадение количества колонок
                                    lst_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=name_sheet)
                                    if lst_df.shape[1] != dct_df[name_sheet].shape[1]:
                                        # если количество колонок не совпадает то записываем как ошибку
                                        temp_error_df = pd.DataFrame(
                                            columns=['Название файла', 'Наименование листа', 'Тип ошибки',
                                                     'Описание ошибки'],
                                            data=[[name_file, name_sheet, 'Количество колонок отличается от эталонного',
                                                   f'Ожидалось {dct_df[name_sheet].shape[1]} колонок, а в листе {lst_df.shape[1]}']])  # создаем временный датафрейм. потом надо подумать над словарем
                                        err_df = pd.concat([err_df, temp_error_df],
                                                           ignore_index=True)  # добавляем в датафрейм ошибок
                                        count_errors += 1
                                #
                                # если хоть одна ошибка то проверяем следующий файл
                                if count_errors != 0:
                                    continue
                                # если нет то начинаем обрабатывать листы
                                for name_sheet, skip_r in dct_manage_harvest.items():
                                    temp_df = pd.read_excel(f'{dirpath}/{filename}', sheet_name=name_sheet,
                                                            skiprows=skip_r,
                                                            dtype=str, header=None)  # загружаем датафрейм

                                    if temp_df.shape[1] > 3:
                                        temp_df = temp_df.dropna(axis=0, thresh=2)
                                    temp_df['Номер строки'] = range(1, temp_df.shape[0] + 1)
                                    temp_df['Откуда взяты данные'] = name_file
                                    for row in dataframe_to_rows(temp_df, index=False, header=False):
                                        standard_wb[name_sheet].append(row)  # добавляем данные
                            else:
                                temp_error_df = pd.DataFrame(
                                    columns=['Название файла', 'Наименование листа', 'Тип ошибки', 'Описание ошибки'],
                                    data=[[name_file, '', 'Не совпадает количество или название листов в файле',
                                           f'Листы, которые есть в файле: {",".join(temb_wb.sheetnames)}']])  # создаем временный датафрейм. потом надо подумать над словарем

                                err_df = pd.concat([err_df, temp_error_df],
                                                   ignore_index=True)  # добавляем в датафрейм ошибок

                # # Получаем текущую дату
                current_time = time.strftime('%H_%M_%S %d.%m.%Y')

                standard_wb.save(
                    f'{path_to_end_folder_merger}/Слияние по варианту В Общая таблица от {current_time}.xlsx')  # сохраняем
                err_out_wb = openpyxl.Workbook()  # создаем объект openpyxl для сохранения датафрейма
                for row in dataframe_to_rows(err_df, index=False, header=True):
                    err_out_wb['Sheet'].append(row)  # добавляем данные
                # устанавливаем размер колонок
                err_out_wb['Sheet'].column_dimensions['A'].width = 40
                err_out_wb['Sheet'].column_dimensions['B'].width = 30
                err_out_wb['Sheet'].column_dimensions['C'].width = 55
                err_out_wb['Sheet'].column_dimensions['D'].width = 100
                err_out_wb.save(f'{path_to_end_folder_merger}/Слияние по варианту В Ошибки от {current_time}.xlsx')

        except NameError:
            messagebox.showerror('Веста Обработка таблиц и создание документов',
                                 f'Выберите папку с файлами,эталонный файл и папку куда будут генерироваться файлы')
        except PermissionError:
            messagebox.showerror('Веста Обработка таблиц и создание документов',
                                 f'Закройте файл выбранный эталонным или файлы из обрабатываемой папки')
        except NotParamsHarvest:
            messagebox.showerror('Веста Обработка таблиц и создание документов',
                                 f'Выберите файл с параметрами объединения таблиц')
        except FileNotFoundError:
            messagebox.showerror('Веста Обработка таблиц и создание документов',
                                 f'Выберите файл с параметрами!\n'
                                 f'Если вы выбрали файл с параметрами, а ошибка повторяется,то перенесите папку \n'
                                 f'с файлами которые вы хотите обработать в корень диска. Проблема может быть в \n '
                                 f'в слишком длинном пути к обрабатываемым файлам или конечной папке')

        except:
            logging.exception('AN ERROR HAS OCCURRED')
            messagebox.showerror('Веста Обработка таблиц и создание документов',
                                 'Возникла ошибка!!! Подробности ошибки в файле error.log')
        else:
            messagebox.showinfo('Веста Обработка таблиц и создание документов',
                                'Создание общей таблицы успешно завершено!!!')

if __name__=='__main__':
    checkbox_harvest_main = 1
    merger_entry_skip_rows_main = 1
    file_standard_merger_main = 'data\Слияние данных\Списки\Список 28.03.02 Наноинженерия.xlsx'
    dir_name_main = 'data\Слияние данных\Списки'
    path_to_end_folder_merger_main = 'data'
    file_params_main = ''



    union_tables(checkbox_harvest_main, merger_entry_skip_rows_main, file_standard_merger_main, dir_name_main, path_to_end_folder_merger_main,
                 file_params_main)

    print('Lindy Booth')