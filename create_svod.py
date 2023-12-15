"""
Скрипт для создания простейших сводных таблиц по выбранным колонкам
"""
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import time
from tkinter import messagebox
import numpy as np
import re
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)

class NotSheet(Exception):
    """
    Исключения для случая когла листа нет в датафрейме
    """
    pass



class NotNumberStr(Exception):
    """
    Исключение для случаев когда в строке нет цифр
    """
    pass

class WrongNumberColumn(Exception):
    """
    Исключения для случаев когда порядковый номер колонки больше чем количество колонок в датафрейме
    """
    pass



def count_uniq(value):
    ser_stat = value.describe()  # плоучаем статистику
    return ser_stat.iloc[1]  # возвращаем количество уникальных значений


def find_top_value(value):
    ser_stat = value.describe()  # плоучаем статистику
    return ser_stat.iloc[2]  # возвращаем самое частое значение


def count_top_value(value):
    ser_stat = value.describe()  # плоучаем статистику
    return ser_stat.iloc[3]  # возвращаем количество самых частых значений


def count_dupl_value(value):
    lst_dupl = [value for value in value.duplicated(keep=False).tolist() if value]
    return len(lst_dupl)  # количество дубликатов


def add_percentage(group,lst_target_name_column):
    total_values = group[lst_target_name_column].sum(axis=1)
    for column in lst_target_name_column:
        group[f'{column}_Percentage'] = (group[column] / total_values) * 100
    return group


def generate_svod_for_columns(file_data:str,sheet_name:str,end_folder:str,str_column:str,str_target_column:str):
    """
    Функция для создания сводных таблиц по выбранных колонкам
    Создаются следующие своды ['Сумма',
                             'Среднее',
                             'Медиана',
                             'Минимум',
                             'Максимум',
                             'Количество',
                             'Количество уникальных',
                             'Самое частое',
                             'Количество самых частых',
                             'Количество дубликатов']
    :param file_data:путь к файлу с данными
    :param sheet_name: название листа где находятся данные
    :param end_folder:конечная папка
    :param str_column: строка с номерами колонок по которым будет идти свод вида 1,23,15
    :param str_target_column: строка с номераки колонок по которым будет происходить подсчет
    :return:файл Excel  в котором будет создано 10 листов по названиям сводов
    """

    try:
        temp_wb = openpyxl.load_workbook(file_data,read_only=True)
        if sheet_name not in temp_wb.sheetnames:
            raise NotSheet
        temp_wb.close()
        base_df = pd.read_excel(file_data, sheet_name=sheet_name)
        base_df.fillna('Не заполнено',inplace=True)

        # обрабатываем список колонок по которым нужно группировать
        _lst_cols = re.findall(r'\d+', str_column)  # находим цифры
        if _lst_cols:
            lst_number_cols = list(map(int, _lst_cols))
        else:
            raise NotNumberStr

        _target_col = re.findall(r'\d+', str_target_column)
        if _target_col:
            lst_target_number_column = list(map(int, _target_col))
        else:
            raise NotNumberStr

        # Проверка на существование колонки
        for value in lst_number_cols:
            if value - 1 >= base_df.shape[1]:
                error_value  = value
                raise WrongNumberColumn

        for value in lst_target_number_column:
            if value - 1 >= base_df.shape[1]:
                error_value = value
                raise WrongNumberColumn

        lst_name_column = []  # список для хранения названий колонок сводной таблицы
        for value in lst_number_cols:
            if value == 0:
                lst_name_column.append(base_df.columns[0])
            else:
                lst_name_column.append(base_df.columns[value - 1])

        lst_target_name_column = []  # список для хранения названий колонок сводной таблицы
        for value in lst_target_number_column:
            if value == 0:  # обработка нуля
                lst_target_name_column.append(base_df.columns[0])
            else:
                lst_target_name_column.append(base_df.columns[value - 1])

        dct_func = {'Сумма': 'sum', 'Среднее': 'mean', 'Медиана': 'median', 'Минимум': 'min', 'Максимум': 'max',
                    'Количество': 'count',
                    'Количество уникальных': count_uniq, 'Самое частое': find_top_value,
                    'Количество самых частых': count_top_value,
                    'Количество дубликатов': count_dupl_value}

        set_text_func = {'Количество уникальных', 'Самое частое', 'Количество самых частых', 'Количество дубликатов'}

        wb = openpyxl.Workbook()  # создаем объект
        for idx, items in enumerate(dct_func.items()):
            name_sheet, name_func = items  # распаковываем
            wb.create_sheet(title=name_sheet, index=idx)  # создаем лист
            df = base_df.copy()  # копируем дотафрейм что избежать косяков с изменением значений
            if name_sheet not in set_text_func:

                df[lst_target_name_column] = df[lst_target_name_column].applymap(
                    lambda x: x if isinstance(x, (int, float)) else 0)
                if name_sheet == 'Количество':
                    value = lst_target_name_column[0]

                    temp_df = pd.pivot_table(df,
                                             index=lst_name_column,
                                             values=value,
                                             aggfunc=name_func,
                                             fill_value=0)
                    # Получаем названия колонок по которым будет проводить группировку внутри группы

                    # заменяем название колонки
                    temp_df.columns = ['Количество']
                    temp_df = temp_df.reset_index() # извлекаем индекс
                    # считаем проценты внутри группы
                    if len(lst_name_column) == 1:
                        all_sum = temp_df['Количество'].sum()  # получаем общее количество
                        # добавляем колонку с процентом от общего количества
                        temp_df['Доля в % от общего количества'] = ((temp_df['Количество'] / all_sum)).apply(
                            lambda x: (round(x, 3)) * 100)
                    else:
                        temp_df['Доля в % внутри группы'] = temp_df['Количество'] / temp_df.groupby(lst_name_column[:-1])['Количество'].transform('sum')
                        # приводим к читаемому виду
                        temp_df['Доля в % внутри группы'] = temp_df['Доля в % внутри группы'].apply(
                            lambda x: (round(x, 3)) * 100)
                        all_sum = temp_df['Количество'].sum() # получаем общее количество
                        # добавляем колонку с процентом от общего количества
                        temp_df['Доля в % от общего количества'] = ((temp_df['Количество'] / all_sum)).apply(
                            lambda x: (round(x, 3)) * 100)

                else:
                    temp_df = pd.pivot_table(df,
                                             index=lst_name_column,
                                             values=lst_target_name_column,
                                             aggfunc=name_func,
                                             fill_value=0)
                    temp_df = temp_df.reset_index()

                for row in dataframe_to_rows(temp_df, index=False, header=True):
                    wb[name_sheet].append(row)
                for column in wb[name_sheet].columns:
                    max_length = 0
                    column_name = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    wb[name_sheet].column_dimensions[column_name].width = adjusted_width
            else:
                df[lst_target_name_column] = df[lst_target_name_column].fillna('Не заполнено')
                df[lst_target_name_column] = df[lst_target_name_column].applymap(str)
                temp_df = pd.pivot_table(df,
                                         index=lst_name_column,
                                         values=lst_target_name_column,
                                         aggfunc=name_func,
                                         )
                temp_df = temp_df.reset_index()

                for row in dataframe_to_rows(temp_df, index=False, header=True):
                    wb[name_sheet].append(row)
                for column in wb[name_sheet].columns:
                    max_length = 0
                    column_name = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    wb[name_sheet].column_dimensions[column_name].width = adjusted_width

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        wb.save(f'{end_folder}/Сводные данные {current_time}.xlsx')
    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
        logging.exception('AN ERROR HAS OCCURRED')
    except NotSheet:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'В файле нет листа с названием: {sheet_name}. Проверьте написание!')
        logging.exception('AN ERROR HAS OCCURRED')

    except NotNumberStr:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Введите порядковые номера колонок в виде цифр разделенных запятыми !')
        logging.exception('AN ERROR HAS OCCURRED')

    except WrongNumberColumn:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'В таблице нет колонки с таким порядковым номером: {error_value} \n'
                             f'В таблице только {base_df.shape[1]} колонок ')
        logging.exception('AN ERROR HAS OCCURRED')
    except KeyError as e:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'В таблице не найдена указанная колонка {e.args}')

    except FileNotFoundError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')
    # except:
    #     logging.exception('AN ERROR HAS OCCURRED')
    #     messagebox.showerror('Веста Обработка таблиц и создание документов',
    #                          'Возникла ошибка!!! Подробности ошибки в файле error.log')
    else:
        messagebox.showinfo('Веста Обработка таблиц и создание документов', 'Данные успешно обработаны')


if __name__ =='__main__':
    file_data_main = 'data/Сводная таблица/Содействие занятости 2023.xlsx'
    # file_data_main = 'data/Сводная таблица/Билет в будущее сводный отчет по ученикам осень 2023.xlsx'
    # sheet_name_main = 'Заявки'
    sheet_name_main = 'Заявки'
    end_folder_main = 'data/Сводная таблица/result'
    str_column_main = '12,7'  # колонки для сводной таблицы
    # str_column_main = 'fgg'
    str_target_column_main = '3'  # целевая колонка

    generate_svod_for_columns(file_data_main,sheet_name_main,end_folder_main,str_column_main,str_target_column_main)

    print('Lindy Booth')

