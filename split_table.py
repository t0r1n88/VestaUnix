"""
Скрипт для разделения списка по значениям выбранной колонки.Результаты сохраняются либо в листы одного файла либо
в отдельные файлы. Например разделить большой список по полу или по группам
"""
import numpy as np
import pandas as pd
import re
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from tkinter import messagebox
import time
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)

class ZeroNumberColumn(Exception):
    """
    Исключение если будет введен ноль в поле ввода номера колонки
    """
    pass


class ExceedingQuantity(Exception):
    """
    Исключение для случаев когда числа уникальных значений больше 255
    """
    pass


def clean_value(value):
    """
    Функция для обработки значений колонки от  пустых пробелов,нан
    :param value: значение ячейки
    :return: очищенное значение
    """
    if value is np.nan:
        return 'Не заполнено'
    str_value = str(value)
    if str_value == '':
        return 'Не заполнено'
    elif str_value ==' ':
        return 'Не заполнено'

    return str_value



def split_table(file_data_split:str,number_column:int,checkbox_split:int,path_to_end_folder):
    """
    Функция для разделения таблицы по значениям в определенном листе и колонке. Разделение по файлам и листам с сохранением названий

    :param file_data_split: файл с таблицей
    :param number_column:порядковый номер колонки , прибавляется 1 чтобы соответстовать экселю
    :param checkbox_split: вариант разделения
    :param path_to_end_folder: путь к итоговой папке
    :return: один файл в котором много листов либо много файлов в зависимости от режима
    """

    try:
        if number_column == 0: # если кто нажал
            raise ZeroNumberColumn
        df = pd.read_excel(file_data_split,dtype=str)
        name_column = df.columns[number_column - 1]  # получаем название колонки
        df[name_column] = df[name_column].apply(clean_value)

        lst_value_column = df.iloc[:,number_column-1].unique() # получаем все значения нужной колонки, -1 отнимаем поскольку в экселе нумерация с 1

        lst_value_column = list(map(str,lst_value_column))
        used_name_sheet = set() # множество для хранения значений которые уже были использованы
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S',t)

        if checkbox_split == 0:
            if len(lst_value_column) >= 253:
                raise ExceedingQuantity
            wb = openpyxl.Workbook() # создаем файл
            for idx,value in enumerate(lst_value_column):
                temp_df = df[df[name_column] == value] # отфильтровываем по значению
                short_value = value[:20] # получаем обрезанное значение
                short_value = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_value)

                if short_value in used_name_sheet:
                    short_value = f'{short_value}_{idx}' # добавляем окончание
                wb.create_sheet(short_value,index=idx) # создаем лист
                used_name_sheet.add(short_value)
                for row in dataframe_to_rows(temp_df,index=False,header=True):
                    wb[short_value].append(row)

                # Устанавливаем автоширину для каждой колонки
                for column in wb[short_value].columns:
                    max_length = 0
                    column_name = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    wb[short_value].column_dimensions[column_name].width = adjusted_width
            wb.save(f'{path_to_end_folder}\Вариант А один файл {current_time}.xlsx')
            wb.close()


        else:
            used_name_file = set() # множество для уже использованных имен файлов
            for idx,value in enumerate(lst_value_column):
                wb = openpyxl.Workbook()  # создаем файл
                temp_df = df[df[name_column] == value] # отфильтровываем по значению
                short_name = value[:40] # получаем обрезанное значение
                short_name = re.sub(r'[\'+()<> :"?*|\\/]', '_', short_name)
                if short_name in used_name_file:
                    short_name = f'{short_name}_{idx}' # добавляем окончание
                for row in dataframe_to_rows(temp_df,index=False,header=True):
                    wb['Sheet'].append(row)

                # Устанавливаем автоширину для каждой колонки
                for column in wb['Sheet'].columns:
                    max_length = 0
                    column_name = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    wb['Sheet'].column_dimensions[column_name].width = adjusted_width

                wb.save(f'{path_to_end_folder}\{short_name}.xlsx')
                used_name_file.add(short_name)
                wb.close()
    except NameError as e:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
        logging.exception('AN ERROR HAS OCCURRED')
    except ValueError as e:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'В таблице не найден указанный лист {e.args}')
    except ZeroNumberColumn:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Порядковые номера колонок начинаются с 1 !!!')
    except ExceedingQuantity:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Количество уникальных значений в выбранной колонке больше 253!!!\n'
                             f'Выберите вариант Б для создания отдельных файлов или уменьшите количество уникальных значений')


    except IndexError as e:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'В таблице нет колонки с таким порядковым номером')
    except PermissionError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Закройте все файлы Word созданные Вестой')
        logging.exception('AN ERROR HAS OCCURRED')
    except FileNotFoundError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')
    else:
        messagebox.showinfo('Веста Обработка таблиц и создание документов', 'Разделение таблицы завершено!')


if __name__ == '__main__':
    file_data = 'data/Разделение таблицы/Базовая таблица 1000 человек.xlsx'
    name_sheet_main = 'Sheet1'
    number_column_main = 16
    checkbox_split_main = 0
    path_to_end_folder_main = 'data/Разделение таблицы/result'

    split_table(file_data,name_sheet_main, number_column_main, checkbox_split_main, path_to_end_folder_main)
    print('Lindy Booth')



